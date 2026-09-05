from __future__ import annotations

import csv
import os
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path

import numpy as np

from .arkit import (
    ARKitFrame,
    invert_transform,
    read_frame_manifest,
    rotation_to_quaternion_wxyz,
)


@dataclass(frozen=True)
class HLocConfig:
    """Configuration for AnyLoc retrieval + HLoc metric localization."""

    image_root: Path
    reference_manifest: Path
    query_manifest: Path
    anyloc_similarity_csv: Path
    output_dir: Path
    hloc_root: Path | None = None
    retrieval_top_k: int = 20
    temporal_window: int = 5
    local_feature_conf: str = "superpoint_aachen"
    matcher_conf: str = "superpoint+lightglue"
    ransac_threshold_px: float = 12.0
    overwrite_features: bool = False
    overwrite_matches: bool = False


@dataclass(frozen=True)
class HLocPose:
    image_name: str
    camera_from_world: np.ndarray
    pnp_success: bool | None = None
    num_inliers: int | None = None

    @property
    def world_from_camera(self) -> np.ndarray:
        return invert_transform(self.camera_from_world)


def _require_openpyxl():
    try:
        from openpyxl import Workbook
    except ImportError as error:
        raise ImportError(
            "Excel export needs openpyxl. Install the localization requirements "
            "with: pip install -r requirements-localization.txt"
        ) from error
    return Workbook


def _split_names(records: list[ARKitFrame], split: str) -> list[str]:
    return [f"{split}/{record.image_name}" for record in records]


def _top_indices(scores: np.ndarray, top_k: int) -> np.ndarray:
    count = min(top_k, len(scores))
    if count == len(scores):
        return np.argsort(-scores)
    candidates = np.argpartition(scores, -count)[-count:]
    return candidates[np.argsort(-scores[candidates])]


def _similarity_rows(path: Path, database_count: int):
    with Path(path).open(newline="") as handle:
        reader = csv.reader(handle)
        header = next(reader, None)
        if header is None:
            raise ValueError(f"AnyLoc similarity CSV is empty: {path}")
        for row_index, row in enumerate(reader):
            scores = np.asarray(row[2:], dtype=np.float32)
            if len(scores) != database_count:
                raise ValueError(
                    f"Similarity row {row_index} contains {len(scores)} scores; "
                    f"expected {database_count}."
                )
            yield row_index, row[1], scores


def write_query_retrieval_pairs(
    similarity_csv: str | Path,
    reference_records: list[ARKitFrame],
    query_records: list[ARKitFrame],
    output_path: str | Path,
    *,
    top_k: int,
) -> Path:
    """Convert the streamed AnyLoc matrix to HLoc query-reference pairs."""

    if top_k < 1:
        raise ValueError("top_k must be at least 1.")
    ref_names = _split_names(reference_records, "reference")
    query_names = _split_names(query_records, "query")
    pairs = []
    for row_index, _, scores in _similarity_rows(Path(similarity_csv), len(ref_names)):
        if row_index >= len(query_names):
            raise ValueError("Similarity CSV has more rows than the query manifest.")
        pairs.extend((query_names[row_index], ref_names[index]) for index in _top_indices(scores, top_k))
    if len(pairs) != len(query_names) * min(top_k, len(ref_names)):
        raise ValueError("Similarity CSV row count does not match the query manifest.")
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(f"{query} {reference}" for query, reference in pairs) + "\n")
    return output_path


def write_reference_temporal_pairs(
    reference_records: list[ARKitFrame],
    output_path: str | Path,
    *,
    temporal_window: int,
) -> Path:
    """Pair each map frame with its next overlapping temporal neighbors."""

    if temporal_window < 1:
        raise ValueError("temporal_window must be at least 1.")
    names = _split_names(reference_records, "reference")
    pairs = [
        (names[index], names[other])
        for index in range(len(names))
        for other in range(index + 1, min(len(names), index + temporal_window + 1))
    ]
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(f"{left} {right}" for left, right in pairs) + "\n")
    return output_path


def write_query_list(query_records: list[ARKitFrame], output_path: str | Path) -> Path:
    """Write HLoc's calibrated query-list format using ARKit intrinsics."""

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        " ".join(
            map(
                str,
                (
                    f"query/{record.image_name}",
                    "PINHOLE",
                    record.width_px,
                    record.height_px,
                    record.fx_px,
                    record.fy_px,
                    record.cx_px,
                    record.cy_px,
                ),
            )
        )
        for record in query_records
    ]
    output_path.write_text("\n".join(lines) + "\n")
    return output_path


def build_known_pose_reference_model(
    reference_records: list[ARKitFrame], output_dir: str | Path
):
    """Create a metric COLMAP model directly from ARKit ``T_world_camera`` poses."""

    try:
        import pycolmap
    except ImportError as error:
        raise ImportError(
            "Known-pose map construction requires pycolmap. Install "
            "requirements-localization.txt."
        ) from error

    reconstruction = pycolmap.Reconstruction()
    for image_id, record in enumerate(reference_records, start=1):
        camera = pycolmap.Camera(
            model="PINHOLE",
            width=record.width_px,
            height=record.height_px,
            params=np.asarray(
                [record.fx_px, record.fy_px, record.cx_px, record.cy_px],
                dtype=np.float64,
            ),
            camera_id=image_id,
        )
        reconstruction.add_camera_with_trivial_rig(camera)
        image = pycolmap.Image(
            name=f"reference/{record.image_name}",
            camera_id=image_id,
            image_id=image_id,
        )
        camera_from_world = record.camera_from_world
        pose = pycolmap.Rigid3d(camera_from_world[:3, :])
        reconstruction.add_image_with_trivial_frame(image, pose)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    reconstruction.write(output_dir)
    return reconstruction


def parse_hloc_results(results_path: str | Path) -> dict[str, HLocPose]:
    poses: dict[str, HLocPose] = {}
    path = Path(results_path)
    success: dict[str, tuple[bool, int | None]] = {}
    status_path = Path(f"{path}_status.csv")
    if status_path.is_file():
        with status_path.open(newline="") as handle:
            for row in csv.DictReader(handle):
                success[row["image_name"]] = (
                    row["pnp_success"].lower() == "true",
                    None if not row["num_inliers"] else int(row["num_inliers"]),
                )
    with path.open() as handle:
        for line in handle:
            if not line.strip():
                continue
            name, *values = line.split()
            qw, qx, qy, qz, tx, ty, tz = map(float, values[:7])
            from .arkit import quaternion_wxyz_to_rotation

            transform = np.eye(4, dtype=np.float64)
            transform[:3, :3] = quaternion_wxyz_to_rotation(
                np.asarray([qw, qx, qy, qz])
            )
            transform[:3, 3] = [tx, ty, tz]
            status = success.get(name, (None, None))
            poses[name] = HLocPose(name, transform, status[0], status[1])
    return poses


def export_hloc_poses_xlsx(results_path: str | Path, output_path: str | Path) -> Path:
    Workbook = _require_openpyxl()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "poses"
    sheet.append(
        [
            "query_image",
            "pnp_success",
            "num_inliers",
            "tx_m",
            "ty_m",
            "tz_m",
            "qw",
            "qx",
            "qy",
            "qz",
            "pose_convention",
        ]
    )
    for pose in parse_hloc_results(results_path).values():
        world_from_camera = pose.world_from_camera
        q = rotation_to_quaternion_wxyz(world_from_camera[:3, :3])
        sheet.append(
            [
                pose.image_name,
                pose.pnp_success,
                pose.num_inliers,
                *world_from_camera[:3, 3].tolist(),
                *q.tolist(),
                "T_reference_world_camera",
            ]
        )
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def export_anyloc_retrieval_xlsx(
    similarity_csv: str | Path,
    reference_manifest: str | Path,
    query_manifest: str | Path,
    output_path: str | Path,
    *,
    top_k: int = 20,
) -> Path:
    """Export top AnyLoc retrievals to Excel without loading the full matrix."""

    Workbook = _require_openpyxl()
    references = read_frame_manifest(reference_manifest)
    queries = read_frame_manifest(query_manifest)
    workbook = Workbook(write_only=True)
    matches = workbook.create_sheet("top_matches")
    matches.append(
        [
            "query_frame",
            "query_image",
            "query_tx_m",
            "query_ty_m",
            "query_tz_m",
            "rank",
            "reference_frame",
            "reference_image",
            "similarity",
            "reference_tx_m",
            "reference_ty_m",
            "reference_tz_m",
        ]
    )
    row_count = 0
    for query_index, _, scores in _similarity_rows(Path(similarity_csv), len(references)):
        query = queries[query_index]
        for rank, reference_index in enumerate(_top_indices(scores, top_k), start=1):
            reference = references[int(reference_index)]
            matches.append(
                [
                    query.frame_index,
                    query.image_name,
                    query.tx_m,
                    query.ty_m,
                    query.tz_m,
                    rank,
                    reference.frame_index,
                    reference.image_name,
                    float(scores[reference_index]),
                    reference.tx_m,
                    reference.ty_m,
                    reference.tz_m,
                ]
            )
            row_count += 1
    config = workbook.create_sheet("summary")
    config.append(["field", "value"])
    config.append(["query_frames", len(queries)])
    config.append(["reference_frames", len(references)])
    config.append(["top_k", min(top_k, len(references))])
    config.append(["match_rows", row_count])
    config.append(["similarity_csv", str(Path(similarity_csv).resolve())])
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def run_hloc(config: HLocConfig) -> dict[str, Path]:
    """Run local-feature map triangulation and 6-DoF query localization."""

    output_dir = Path(config.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    references = read_frame_manifest(config.reference_manifest)
    queries = read_frame_manifest(config.query_manifest)
    all_names = _split_names(references, "reference") + _split_names(queries, "query")
    all_images = output_dir / "all-images.txt"
    all_images.write_text("\n".join(all_names) + "\n")

    reference_pairs = write_reference_temporal_pairs(
        references, output_dir / "pairs-reference-temporal.txt", temporal_window=config.temporal_window
    )
    query_pairs = write_query_retrieval_pairs(
        config.anyloc_similarity_csv,
        references,
        queries,
        output_dir / "pairs-query-anyloc.txt",
        top_k=config.retrieval_top_k,
    )
    query_list = write_query_list(queries, output_dir / "queries-with-intrinsics.txt")
    known_model = output_dir / "reference_arkit_model"
    feature_outputs = {
        "superpoint_aachen": "feats-superpoint-n4096-r1024",
        "superpoint_max": "feats-superpoint-n4096-rmax1600",
        "superpoint_inloc": "feats-superpoint-n4096-r1600",
        "disk": "feats-disk",
        "aliked-n16": "feats-aliked-n16",
    }
    matcher_outputs = {
        "superpoint+lightglue": "matches-superpoint-lightglue",
        "disk+lightglue": "matches-disk-lightglue",
        "aliked+lightglue": "matches-aliked-lightglue",
        "superglue": "matches-superglue",
        "superglue-fast": "matches-superglue-it5",
    }
    if config.local_feature_conf not in feature_outputs:
        raise ValueError(f"Unsupported local_feature_conf: {config.local_feature_conf}")
    if config.matcher_conf not in matcher_outputs:
        raise ValueError(f"Unsupported matcher_conf: {config.matcher_conf}")
    features = output_dir / f"{feature_outputs[config.local_feature_conf]}.h5"
    matches = output_dir / f"{matcher_outputs[config.matcher_conf]}.h5"

    source_root = Path(__file__).resolve().parents[2]
    environment = os.environ.copy()
    python_paths = [str(source_root)]
    if config.hloc_root is not None:
        hloc_root = Path(config.hloc_root).expanduser().resolve()
        if not (hloc_root / "hloc").is_dir():
            raise FileNotFoundError(f"HLoc package not found under: {hloc_root}")
        python_paths.append(str(hloc_root))
    if environment.get("PYTHONPATH"):
        python_paths.append(environment["PYTHONPATH"])
    environment["PYTHONPATH"] = os.pathsep.join(python_paths)

    def stage(name: str, *arguments: object) -> None:
        command = [sys.executable, "-m", "indoor_vpr.localization._hloc_stage", name]
        command.extend(str(argument) for argument in arguments)
        subprocess.run(command, check=True, env=environment)

    stage("known-model", "--manifest", config.reference_manifest, "--output", known_model)
    extract_args: list[object] = [
        "--conf", config.local_feature_conf,
        "--image-root", config.image_root,
        "--image-list", all_images,
        "--features", features,
    ]
    if config.overwrite_features:
        extract_args.append("--overwrite")
    stage("extract", *extract_args)
    for pairs_path in (reference_pairs, query_pairs):
        match_args: list[object] = [
            "--conf", config.matcher_conf,
            "--pairs", pairs_path,
            "--features", features,
            "--matches", matches,
        ]
        if config.overwrite_matches:
            match_args.append("--overwrite")
        stage("match", *match_args)

    reference_sfm = output_dir / "reference_sfm"
    stage(
        "triangulate",
        "--output", reference_sfm,
        "--reference-model", known_model,
        "--image-root", config.image_root,
        "--pairs", reference_pairs,
        "--features", features,
        "--matches", matches,
    )
    results = output_dir / "hloc_poses.txt"
    stage(
        "localize",
        "--reference-sfm", reference_sfm,
        "--queries", query_list,
        "--pairs", query_pairs,
        "--features", features,
        "--matches", matches,
        "--results", results,
        "--ransac-threshold", config.ransac_threshold_px,
    )
    poses_xlsx = export_hloc_poses_xlsx(results, output_dir / "hloc_poses.xlsx")
    return {
        "reference_pairs": reference_pairs,
        "query_pairs": query_pairs,
        "query_list": query_list,
        "features": features,
        "matches": matches,
        "reference_sfm": reference_sfm,
        "poses_txt": results,
        "poses_xlsx": poses_xlsx,
    }
