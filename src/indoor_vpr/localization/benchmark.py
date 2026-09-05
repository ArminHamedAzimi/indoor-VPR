from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np

from .arkit import (
    ARKitFrame,
    first_pose_alignment,
    read_frame_manifest,
    rotation_error_degrees,
)
from .hloc_pipeline import HLocPose, parse_hloc_results


@dataclass(frozen=True)
class FrameError:
    frame_index: int
    image_name: str
    pnp_success: bool | None
    num_inliers: int | None
    translation_error_m: float | None
    rotation_error_deg: float | None
    estimated_world_from_camera: np.ndarray | None
    ground_truth_world_from_camera: np.ndarray


DEFAULT_THRESHOLDS = ((0.25, 2.0), (0.5, 5.0), (1.0, 10.0))


def _errors_for_alignment(
    queries: list[ARKitFrame],
    poses: dict[str, HLocPose],
    reference_world_from_query_world: np.ndarray,
) -> list[FrameError]:
    errors = []
    for query in queries:
        name = f"query/{query.image_name}"
        estimate = poses.get(name)
        gt = reference_world_from_query_world @ query.world_from_camera
        usable = estimate is not None and estimate.pnp_success is not False
        if usable:
            estimated = estimate.world_from_camera
            translation_error = float(np.linalg.norm(estimated[:3, 3] - gt[:3, 3]))
            rotation_error = rotation_error_degrees(estimated[:3, :3], gt[:3, :3])
        else:
            estimated = None
            translation_error = None
            rotation_error = None
        errors.append(
            FrameError(
                frame_index=query.frame_index,
                image_name=name,
                pnp_success=None if estimate is None else estimate.pnp_success,
                num_inliers=None if estimate is None else estimate.num_inliers,
                translation_error_m=translation_error,
                rotation_error_deg=rotation_error,
                estimated_world_from_camera=estimated,
                ground_truth_world_from_camera=gt,
            )
        )
    return errors


def _summarize(
    errors: list[FrameError], thresholds: tuple[tuple[float, float], ...]
) -> dict[str, float | int]:
    valid = [error for error in errors if error.translation_error_m is not None]
    translation = np.asarray([error.translation_error_m for error in valid], dtype=float)
    rotation = np.asarray([error.rotation_error_deg for error in valid], dtype=float)
    total = len(errors)
    summary: dict[str, float | int] = {
        "query_count": total,
        "pnp_success_count": len(valid),
        "pnp_success_rate": len(valid) / total if total else 0.0,
        "median_translation_m": float(np.median(translation)) if len(valid) else float("nan"),
        "mean_translation_m": float(np.mean(translation)) if len(valid) else float("nan"),
        "p90_translation_m": float(np.percentile(translation, 90)) if len(valid) else float("nan"),
        "median_rotation_deg": float(np.median(rotation)) if len(valid) else float("nan"),
        "mean_rotation_deg": float(np.mean(rotation)) if len(valid) else float("nan"),
        "p90_rotation_deg": float(np.percentile(rotation, 90)) if len(valid) else float("nan"),
    }
    for translation_limit, rotation_limit in thresholds:
        count = sum(
            error.translation_error_m is not None
            and error.translation_error_m <= translation_limit
            and error.rotation_error_deg <= rotation_limit
            for error in errors
        )
        key = f"recall_{translation_limit:g}m_{rotation_limit:g}deg"
        summary[key] = count / total if total else 0.0
    return summary


def benchmark_hloc(
    results_path: str | Path,
    reference_manifest: str | Path,
    query_manifest: str | Path,
    *,
    thresholds: tuple[tuple[float, float], ...] = DEFAULT_THRESHOLDS,
) -> dict[str, dict[str, object]]:
    """Evaluate HLoc in raw ARKit coordinates and with first-pose alignment.

    The first-pose result assumes both recordings began from the same physical
    camera pose. It is a coordinate-frame alignment, not an HLoc calibration.
    """

    references = read_frame_manifest(reference_manifest)
    queries = read_frame_manifest(query_manifest)
    poses = parse_hloc_results(results_path)
    identity = np.eye(4, dtype=np.float64)
    first_pose = first_pose_alignment(
        references[0].world_from_camera, queries[0].world_from_camera
    )
    benchmark = {}
    for name, transform in (("raw", identity), ("first_pose_aligned", first_pose)):
        errors = _errors_for_alignment(queries, poses, transform)
        benchmark[name] = {
            "errors": errors,
            "summary": _summarize(errors, thresholds),
            "reference_world_from_query_world": transform,
        }
    return benchmark


def export_hloc_benchmark_xlsx(
    results_path: str | Path,
    reference_manifest: str | Path,
    query_manifest: str | Path,
    output_path: str | Path,
    *,
    thresholds: tuple[tuple[float, float], ...] = DEFAULT_THRESHOLDS,
) -> tuple[Path, dict[str, dict[str, object]]]:
    try:
        from openpyxl import Workbook
    except ImportError as error:
        raise ImportError(
            "Excel export needs openpyxl. Install requirements-localization.txt."
        ) from error

    benchmark = benchmark_hloc(
        results_path, reference_manifest, query_manifest, thresholds=thresholds
    )
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    metric_names = list(benchmark["raw"]["summary"].keys())
    summary_sheet.append(["alignment_mode", *metric_names])
    for mode, result in benchmark.items():
        summary = result["summary"]
        summary_sheet.append([mode, *(summary[name] for name in metric_names)])

    for mode, result in benchmark.items():
        sheet = workbook.create_sheet(mode)
        sheet.append(
            [
                "query_frame",
                "query_image",
                "pnp_success",
                "num_inliers",
                "translation_error_m",
                "rotation_error_deg",
                "estimated_tx_m",
                "estimated_ty_m",
                "estimated_tz_m",
                "ground_truth_tx_m",
                "ground_truth_ty_m",
                "ground_truth_tz_m",
            ]
        )
        for error in result["errors"]:
            estimated_translation = (
                [None, None, None]
                if error.estimated_world_from_camera is None
                else error.estimated_world_from_camera[:3, 3].tolist()
            )
            sheet.append(
                [
                    error.frame_index,
                    error.image_name,
                    error.pnp_success,
                    error.num_inliers,
                    error.translation_error_m,
                    error.rotation_error_deg,
                    *estimated_translation,
                    *error.ground_truth_world_from_camera[:3, 3].tolist(),
                ]
            )

    notes = workbook.create_sheet("notes")
    notes.append(["item", "explanation"])
    notes.append(
        [
            "pose convention",
            "Estimated and ground-truth poses are T_reference_world_camera (camera-to-world).",
        ]
    )
    notes.append(
        [
            "raw",
            "Compares both ARKit world frames directly; valid only if the sessions share an origin.",
        ]
    )
    notes.append(
        [
            "first_pose_aligned",
            "Aligns the first query ARKit camera pose to the first reference pose; assumes the same physical start pose.",
        ]
    )
    notes.append(
        [
            "failed PnP",
            "A query without a successful HLoc PnP estimate counts as a failure for recall and has blank errors.",
        ]
    )
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path, benchmark
